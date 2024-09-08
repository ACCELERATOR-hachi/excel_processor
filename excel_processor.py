import re
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import os

# 假设以下函数是处理 Excel 文件的函数
from openpyxl.utils import get_column_letter


# 提取时间
def extract_date(cell_content):
    if cell_content is None:
        return datetime(2100, 1, 1).date()  # 返回一个默认日期2100/1/1
    date_patterns = [
        '(\\d{4}-\\d{2}-\\d{2})',  # 匹配数字日期
        '(\\d{4}/\\d{2}/\\d{2})',  # 添加斜杠分隔的日期格式
        '(\\d{2}/\\d{2}/\\d{4})'  # 添加斜杠分隔的日期格式（月/日/年）
    ]
    for pattern in date_patterns:
        match = re.search(pattern, cell_content)
        if match:
            try:
                date_obj = datetime.strptime(match.group(1), "%Y-%m-%d").date()
                return date_obj
            except ValueError:
                pass
    return None


# 计算两单元格时间差
def calculate_date_difference(cell1, cell2, sheet):
    # 从单元格中提取日期字符串
    date1 = extract_date(sheet[cell1].value)
    date2 = extract_date(sheet[cell2].value)

    # 检查日期字符串是否为None
    if date1 is None or date2 is None:
        print("至少有一个单元格不包含有效的日期。")
        return

    # 尝试将字符串转换为datetime对象
    try:
        datetime1 = datetime.strptime(str(date1), "%Y-%m-%d")
        datetime2 = datetime.strptime(str(date2), "%Y-%m-%d")
    except ValueError:
        print("日期格式不正确，请使用 'YYYY-MM-DD' 格式。")
        return

    # 计算两个日期之间的差异
    date_difference = datetime2 - datetime1
    return date_difference


def calculate_duration(updated_sheet, dur_start_col, dur_end_col, dur_target_col):
    for row_idx, row in enumerate(updated_sheet.iter_rows(), start=2):
        # 使用字符串格式化来构造单元格地址
        start_cell = f"{get_column_letter(dur_start_col)}{row_idx}"
        end_cell = f"{get_column_letter(dur_end_col)}{row_idx}"

        # 计算持续时间
        duration = calculate_date_difference(start_cell, end_cell, updated_sheet)

        # 调试信息
        print(f"Row {row_idx}: Start={start_cell}, End={end_cell}, Duration={duration.days}")

        # 在dur_target_col列添加持续时长
        duration_cell = f"{get_column_letter(dur_target_col)}{row_idx}"
        if duration >= timedelta(days=0):
            updated_sheet[duration_cell].value = f"{duration.days}天"
        elif duration < timedelta(days=0):
            updated_sheet[duration_cell].value = "date error"


def add_timestamp_to_updated_cells(original_file, updated_file, output_file):
    # 加载原版本和更新版本的工作簿
    original_wb = load_workbook(original_file)
    updated_wb = load_workbook(updated_file)

    # 获取两个工作簿的活跃工作表
    original_sheet = original_wb.active
    updated_sheet = updated_wb.active

    # 创建一个新工作簿用于输出
    output_wb = Workbook()
    output_sheet = output_wb.active

    # 遍历更新版本工作表中的行
    for row_idx, row in enumerate(updated_sheet.iter_rows(), start=1):
        # 遍历行中的单元格
        for col_idx, cell in enumerate(row, start=1):
            # 获取原版本中相同位置的单元格
            original_cell = original_sheet.cell(row=row_idx, column=col_idx)

            # 检查单元格的值是否发生了变化
            if cell.value != original_cell.value:
                # 如果单元格值有变化，则添加时间戳
                timestamp = datetime.now().strftime("%Y-%m-%d")
                cell.value = f"{cell.value} ({timestamp})"

        # 如果是第一行，复制标题行
        if row_idx == 1:
            output_sheet.append([cell.value for cell in row])
        else:
            output_sheet.append([cell.value for cell in row])

    # 保存输出工作簿
    output_wb.save(output_file)


def main():
    """
    主函数，包含处理文件和保存结果的逻辑
    """
    global ORIGINAL_FILE, UPDATED_FILE, OUTPUT_FILE
    ORIGINAL_FILE = original_entry.get()
    UPDATED_FILE = updated_entry.get()
    OUTPUT_FILE = output_entry.get()

    if not all([ORIGINAL_FILE, UPDATED_FILE, OUTPUT_FILE]):
        messagebox.showerror("错误", "所有文件路径不能为空。")
        return

    try:
        # 处理文件逻辑...
        add_timestamp_to_updated_cells(ORIGINAL_FILE, UPDATED_FILE, OUTPUT_FILE)
        updated_wb = load_workbook(OUTPUT_FILE)
        updated_sheet = updated_wb.active
        #计算周期#####
        calculate_duration(updated_sheet, dur_start_col=3, dur_end_col=4, dur_target_col=7)
        # 保存工作簿
        updated_wb.save(OUTPUT_FILE)
        messagebox.showinfo("成功", "处理完成，输出文件已保存到: " + OUTPUT_FILE)
    except Exception as e:
        messagebox.showerror("错误", str(e))


# 创建主窗口
root = tk.Tk()
root.title("Excel 文件处理工具")

# 创建标签和输入框
tk.Label(root, text="原始文件路径:").grid(row=0, column=0)
tk.Label(root, text="更新文件路径:").grid(row=1, column=0)
tk.Label(root, text="输出文件路径:").grid(row=2, column=0)

original_entry = tk.Entry(root)
updated_entry = tk.Entry(root)
output_entry = tk.Entry(root)

original_entry.grid(row=0, column=1)
updated_entry.grid(row=1, column=1)
output_entry.grid(row=2, column=1)


# 创建选择文件路径的函数
def choose_file(prompt, entry_var):
    file_path = filedialog.askopenfilename(title=prompt)
    if file_path:
        entry_var.delete(0, tk.END)
        entry_var.insert(0, file_path)


# 创建选择输出文件路径的函数
def choose_output_file():
    # 获取原始文件和更新文件的目录
    original_dir = os.path.dirname(original_entry.get())
    updated_dir = os.path.dirname(updated_entry.get())
    common_dir = os.path.commonpath([original_dir, updated_dir])

    # 设置默认保存路径为原始文件和更新文件的共同目录
    default_path = os.path.join(common_dir, "output.xlsx")

    output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                    title="保存输出文件",
                                                    filetypes=[("Excel 文件", "*.xlsx")],
                                                    initialfile=default_path)
    if output_file_path:
        output_entry.delete(0, tk.END)
        output_entry.insert(0, output_file_path)


# 创建运行脚本的函数
def run_script():
    main()  # 调用 main 函数
    root.destroy()  # 关闭主窗口

# 创建按钮
tk.Button(root, text="选择原始文件", command=lambda: choose_file("选择原始文件:", original_entry)).grid(row=0, column=2)
tk.Button(root, text="选择更新文件", command=lambda: choose_file("选择更新文件:", updated_entry)).grid(row=1, column=2)
tk.Button(root, text="选择输出文件位置", command=choose_output_file).grid(row=2, column=2)
tk.Button(root, text="运行脚本", command=run_script).grid(row=3, column=1)

# 启动事件循环
root.mainloop()
